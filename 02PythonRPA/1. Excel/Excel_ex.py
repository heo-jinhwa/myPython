# pip install pandas
# pip install openpyxl

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, Color, PatternFill


class ExcelRPA:
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)

    # 특정 Cell들의 font 색변경
    def Change_FontColor(self, sheetname, ary_cells):
        ws = self.wb[sheetname]
        f = Font(color='0000FF')
        
        for cell in ary_cells:
            ws[cell].font = f
        
        self.wb.save(self.filename)

    def Merge_Excel(self, targetfile, newfilename):
        df_Source = pd.read_excel(self.filename)
        df_Source.set_index('date', inplace=True)

        df_Target = pd.read_excel(targetfile)
        df_Target.set_index('date', inplace=True)

        df_Merge = pd.DataFrame()
        df_Merge['Source'] = df_Source['total']
        df_Merge['Target'] = df_Target['total']

        df_Merge.to_excel(newfilename)
    
    def Calculation(self, sheetname):
        ws = self.wb[sheetname]
        sum_result = sum([row[0].value for row in ws['B2':'B13']])
        ws['B14'].value = sum_result # ws['B14'].value = 'SUM(B2:B13)'
    
    def Style(self, sheetname, cell):
        # 폰트 꼴, 크기, 두께
        ws = self.wb[sheetname]
        cell_num = ws[cell]
        my_font = Font(name='맑은 고딕', size=15, bold=True)

        # 정렬
        align_center = Alignment(horizontal='center', vertical='center')

        # 테두리 설정
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # 셀 색상 지정
        fill_orange = PatternFill(patternType='solid', fgColor=Color('FFC000'))

        cell_num.font = my_font
        cell_num.alignment = align_center
        cell_num.border = border_thin
        cell_num.fill = fill_orange

    def LoopRange(self, sheetname, cell_range):
        ws = self.wb[sheetname]
        
        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = '0.00'





if __name__ == '__main__':
    EA = ExcelRPA("C:/Users/jinhwaheo/Desktop/PATCH/test.xlsx")
    EA.Change_FontColor("kkk", ["A1", "A2", "B5", "C3"])