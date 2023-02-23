import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.chart import Reference, BarChart, LineChart
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.drawing.image import Image
from random import *

def Create_file():
    wb = Workbook() # 새 워크북 생성
    ws = wb.active # 현재 활성화 된 sheet 가져옴 (첫번째 시트)
    ws.title = "NadoSheet" # sheet의 이름을 변경
    wb.save("sample.xlsx")
    wb.close()

def Sheet():
    wb = Workbook() # 새 워크북 생성
    ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
    ws.title = "1Sheet" # sheet 이름 변경
    ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경
    ws1 = wb.create_sheet("2Sheet") # 주어진 이름으로 Sheet 생성
    ws2 = wb.create_sheet("3Sheet", 2) # 두번째 index에 Sheet 생성

    new_ws = wb["3Sheet"] # Dict 형태로 Sheet에 접근가능
    print(wb.sheetnames) # 모든 Sheet 이름 확인

    target = wb.copy_worksheet(new_ws) # Sheet 복사
    target.title = "CopiedSheet"
    
    wb.save("sample.xlsx")
    wb.close()

def Cell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cell_ex"

    ws["A1"] = 1 # A1셀에 1 입력
    ws["A2"] = 2 
    ws["A3"] = 3 

    ws["B1"] = 4
    ws["B2"] = 5
    ws["B3"] = 6

    print(ws["A1"]) # A1 셀의 정보를 출력 (값을 출력하려면 value 해줘야함)
    print(ws["A1"].value) # A1 셀의 값 출력
    print(ws["A10"].value) # 값이 없을때는 None 출력

    # row = 1, 2, 3
    # column = A, B, C
    print(ws.cell(column=1, row=1).value) # ws["A1"].value랑 같은 의미
    print(ws.cell(column=2, row=1).value) # ws["B1"].value랑 같은 의미
    c = ws.cell(column=3, row=1, value=10) # ws["C1"] = 10
    print(c.value) # ws["C1"]

    # 반복문을 이용해서 랜덤 숫자 채우기
    index = 1
    for x in range(1, 11):
        for y in range(1, 11):
            ws.cell(row=x, column=y, value=index)
            index += 1

    wb.save("sample.xlsx")

def Open_file():
    wb = load_workbook("sample.xlsx") # sample.xlsx 파일에서 wb를 불러옴
    ws = wb.active # 활성화된 sheet

    # cell 불러오기
    for x in range(1, 11):
        for y in range(1, 11):
            print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4. ..
        print()
        
    # cell 개수 모를때
    for x in range(1, ws.max_row+1):
        for y in range(1, ws.max_column+1):
            print(ws.cell(row=x, column=y).value, end=" ") 
        print()

def Cell_range():
    wb = Workbook()
    ws = wb.active

    # 한줄씩 데이터 넣기
    ws.append(["번호", "영어", "수학"])
    for i in range(1, 11): # 10줄 넣기
        ws.append([i, randint(0, 100), randint(0, 100)])

    col_B = ws["B"] # 영어 column만 가져오기
    for cell in col_B:
        print(cell.value)

    col_range = ws["B:C"] # 영어, 수학 column 함께 가져오기
    for cols in col_range:
        for cell in cols:
            print(cell.value)

    row_title = ws[1] # 첫번째 row만 가져오기
    for cell in row_title:
        print(cell.value)

    row_range = ws[2:6] # 1번째 줄인 title을 제외하고 2번째 줄에서 6번째 줄까지 가져오기 (6포함)
    for rows in row_range:
        for cell in rows:
            print(cell.value, end=" ")
        print()

    row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
    for rows in row_range:
        for cell in rows:
            # print(cell.value, end=" ")
            # print(cell.coordinate, end=" ") # 셀의 좌표 정보를 가져오기 
            xy = coordinate_from_string(cell.coordinate)
            # print(xy, end=" ")
            print(xy[0], end="") # A
            print(xy[1], end=" ") # 1
        print()

    # rows vs iter_rows : rows는 모든 row를 가져오고, iter_rows는 범위를 지정해서 가져올 수 있음
    # 전체 rows
    # print(tuple(ws.rows))
    for row in tuple(ws.rows):
        print(row[1].value)

    print()
    # 전체 cols
    # print(tuple(ws.columns))
    for column in tuple(ws.columns):
        print(column[0].value)
    print()

    for row in ws.iter_rows(): # 전체 row
        print(row[2].value)
    print()

    for column in ws.iter_cols(): # 전체 col
        print(column[0].value)
    print()

    for row in ws.iter_rows(min_row=1, max_row=11, min_col=2, max_col=3): # 전체 1번째 줄부터 5번째 줄까지, 2번재 열부터 3번째 열까지 범위 지정 가능
        print(row[0].value, row[1].value) # 영어, 수학
    print()

    for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
        print(col)
    print()

    wb.save("sample.xlsx")

def Search():
    wb = load_workbook("sample.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        # 번호, 영어, 수학
        if int(row[1].value) >= 80:
            print(row[0].value, "번 학생은 영어 천재")

    for row in ws.iter_rows(max_row=1):
        for cell in row:
            if cell.value == "영어":
                cell.value = "컴퓨터"

def Insert():
    wb = load_workbook("sample.xlsx")
    ws = wb.active

    # ws.insert_rows(8) # 8번째 줄이 비워짐
    ws.insert_rows(8 ,5) # 8번째 줄부터 5줄 비워짐

    # ws.insert_cols(2) # B번째 열이 비워짐 (새로운 빈 열이 추가)
    ws.insert_cols(2, 3) # B번째 열부터 3줄 비워짐


    wb.save("sample_insert_rows_cols.xlsx")

def Delete():
    wb = load_workbook("sample.xlsx")
    ws = wb.active

    # ws. delete_rows(8) # 8번째 줄에 있는 7번 학생 데이터 삭제
    # ws.delete_rows(8, 3) # 8번째 줄부터 3개 행 삭제

    # ws. delete_cols(2) # 2번째 열 (B) 삭제
    ws.delete_cols(2, 2) # 2번째 열 (B) 부터 2개 열 삭제

    wb.save("sample_delete_cols.xlsx")

def Move():
    wb = load_workbook("sample.xlsx")
    ws = wb.active
    # # 번호영어 수학
    # # 번호 (국어) 영어 수학
    # ws.move_range("B1:C11", rows=0, cols=1)
    # ws["B1"].value = "국어" # B1셀에 '국어' 입력

    # 번호 영어 수학
    ws.move_range("C1:C11", rows=5, cols=-1)
    wb.save("sample_korean.xlsx")

def Chart():
    wb = load_workbook("sample.xlsx")
    ws = wb.active

    bar_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
    bar_chart = BarChart() # 차트 종류 설정 (Bar, Line, Pie, ...)
    bar_chart.add_data(bar_value) # 차트 데이터 추가
    ws.add_chart(bar_chart, 'E1') # 차트 넣을 위치 정의

    line_value = Reference(ws, min_row=1, max_row=11, min_col=2, max_col=3) # 제목 포함 B1:C11까지의 데이터
    line_chart = LineChart() # 차트 종류 설정 (Bar, Line, Pie, ...)
    line_chart.add_data(line_value, titles_from_data=True) # 차트 데이터 추가 / 계열 -> 영어, 수학 (제목에서 추가 가능)
    line_chart.title = "성적표" # 제목
    line_chart.style = 10 # 미리 정의된 스타일 적용, 사용자 개별 적용 가능
    line_chart.y_axis.title = "점수"
    line_chart.x_axis.title = "번호"

    ws.add_chart(line_chart, 'E20') # 차트 넣을 위치 정의

    wb.save("sample_chart.xlsx")

def CellStyle():
    wb = load_workbook("sample.xlsx")
    ws = wb.active

    # 번호, 영어, 수학
    a1 = ws["A1"] # 번호
    b1 = ws["B1"] # 영어
    c1 = ws["C1"] # 수학

    # A열의 너비를 5로 설정
    ws.column_dimensions["A"].width = 5

    # 1행의 높이를 50으로 설정
    ws.row_dimensions[1].height = 50

    # 스타일 적용
    a1.font = Font(color="FF0000", italic=True, bold=True) # 글자 색상 빨갛게 , 이탤릭, 두껍게
    b1.font = Font(color="CC33FF", name="Arial", strike=True) # 글자 색상 보라색, 폰트 설정, 취소선
    c1.font = Font(color="0000FF", size=20, underline="single") # 글자 색상 파란색, 크기 20, 밑줄

    # 테두리 적용
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    a1.border = thin_border
    b1.border = thin_border
    c1.border = thin_border

    # 배경색 변경 (90점 넘는 셀에 대해서 초록색으로 적용)
    for row in ws.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center") # center, left, right, top, bottom
            if cell.column == 1: # A 번호열은 제외
                continue
            # Cell이 정수형 데이터이고, 90점 보다 높으면
            if isinstance(cell.value, int) and cell.value >= 90:
                cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로
                cell.font = Font(color="FF0000")

    # 틀 고정
    ws.freeze_panes = "B2" # B2 기준으로 틀 고정

    wb.save("sample_style.xlsx")

def Formula():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = datetime.datetime.today() # 오늘 날짜 정보
    ws["A2"] = "=SUM(1, 2, 3)"
    ws["A3"] = "=AVERAGE(1, 2, 3)"

    ws["A4"] = 10
    ws["A5"] = 20
    ws["A6"] = "=SUM(A4:A5)"

    wb.save("sample_formula.xlsx")

def FormulaDataOnly():
    wb = load_workbook("sample_formula.xlsx")
    ws = wb.active

    # 수식 그대로 가져오기
    for row in ws.values:
        for cell in row:
            print(cell)
    
    wb = load_workbook("sample_formula.xlsx", data_only=True)
    ws = wb.active

    # 수식이 아닌 실제 데이터를 가지고 옴
    # evaluate 되지 않은 상태의 데이터는 None이라고 표시
    for row in ws.values:
        for cell in row:
            print(cell)

def Merge():
    wb = Workbook()
    ws = wb.active

    # 병합하기
    ws.merge_cells("B2:D2") # B2~D2까지 합치겠음
    ws["B2"].value = "Merged Cell"
    
    wb.save("sample_merge.xlsx")

def UnMerge():
    wb = load_workbook("sample_merge.xlsx")
    ws = wb.active

    # B2:D2 병합 셀 해제
    ws.unmerge_cells("B2:D2")

    wb.save("sample_unmerge.xlsx")

def ImageEX():
    wb = Workbook()
    ws = wb.active

    img = Image("img.png")
    
    # C3 위치에 이미지 삽입
    ws.add_image(img, "C3")
    wb.save("sample_img.xlsx")

if __name__ == '__main__':
    # Create_file() # 1번 실행
    # Sheet() # 2번 실행
    # Cell() # 3번 실행
    # Open_file() # 4번 실행
    # Cell_range() # 5번 실행
    # Search() # 6번 실행
    # Insert() # 7번 실행
    # Delete() # 8번 실행
    # Move()
    # Chart()
    # CellStyle()
    # Formula()
    # FormulaDataOnly()
    # UnMerge()
    ImageEX()