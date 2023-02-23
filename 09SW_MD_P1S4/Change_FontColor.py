# pip install openpyxl

from openpyxl import load_workbook
from openpyxl.styles import Font 

def Change_FontColor(filename, sheetname, ary_cells):
    wb = load_workbook(filename=filename, keep_vba=True)
    ws = wb[sheetname]
    f = Font(color='0000FF', bold=True)
    
    for cell in ary_cells:
        ws[cell].font = f
    
    wb.save(filename)

if __name__ == '__main__':
    Change_FontColor("../통합 문서1.xlsm", "Sheet1", ["A1", "A2", "B5", "C3", "Y19"])